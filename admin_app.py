import ttkbootstrap as ttk
from ttkbootstrap.constants import *
from tkinter import messagebox, scrolledtext, filedialog
import tkinter as tk
import requests
import json
import string
import random
import threading
import configparser
import os
from datetime import datetime

# 读取配置文件
def load_config():
    """从配置文件加载设置"""
    config = configparser.ConfigParser()
    config_file = 'config.ini'
    
    # 如果配置文件不存在，创建默认配置
    if not os.path.exists(config_file):
        config['API'] = {
            'base_url': 'http://localhost:5000'
        }
        config['APP'] = {
            'default_theme': 'cosmo',
            'window_width': '1000',
            'window_height': '750'
        }
        with open(config_file, 'w', encoding='utf-8') as f:
            config.write(f)
        print(f"已创建默认配置文件: {config_file}")
    else:
        config.read(config_file, encoding='utf-8')
    
    return config

# 加载配置
CONFIG = load_config()
BASE_URL = CONFIG.get('API', 'base_url', fallback='http://localhost:5000')

# 可选主题: 'cosmo', 'flatly', 'litera', 'minty', 'lumen', 'sandstone', 
#          'yeti', 'pulse', 'united', 'morph', 'journal', 'darkly', 
#          'superhero', 'solar', 'cyborg', 'vapor', 'simplex', 'cerculean'

class ModernAdminApp:
    def __init__(self, root):
        self.root = root
        self.root.title("JYH遥控器")
        
        # 创建主框架
        main_frame = ttk.Frame(root, padding=15, bootstyle="secondary")
        main_frame.pack(fill=BOTH, expand=YES)
        
        # 标题区域
        header_frame = ttk.Frame(main_frame)
        header_frame.pack(fill=X, pady=(0, 15))
        
        # 标题
        title_label = ttk.Label(
            header_frame, 
            text="🎮 JYH遥控器", 
            font=("Microsoft YaHei UI", 24, "bold"),
            bootstyle="inverse-primary"
        )
        title_label.pack(side=LEFT, padx=10)
        
        # 主题切换按钮
        theme_frame = ttk.Frame(header_frame)
        theme_frame.pack(side=RIGHT, padx=10)
        
        ttk.Label(theme_frame, text="主题:", font=("Microsoft YaHei UI", 10)).pack(side=LEFT, padx=5)
        
        self.theme_var = ttk.StringVar(value=root.style.theme.name)
        theme_combo = ttk.Combobox(
            theme_frame, 
            textvariable=self.theme_var,
            values=['cosmo', 'flatly', 'litera', 'minty', 'lumen', 'sandstone', 
                   'yeti', 'pulse', 'united', 'morph', 'journal', 'darkly', 
                   'superhero', 'solar', 'cyborg', 'vapor', 'simplex', 'cerculean'],
            state='readonly',
            width=12,
            bootstyle="primary"
        )
        theme_combo.pack(side=LEFT, padx=5)
        theme_combo.bind('<<ComboboxSelected>>', self.change_theme)
        
        # 创建Notebook（选项卡）
        self.notebook = ttk.Notebook(main_frame, bootstyle="primary")
        self.notebook.pack(fill=BOTH, expand=YES, pady=10)
        
        # 创建三个选项卡
        self.create_custom_add_tab()
        self.create_batch_add_tab()
        self.create_query_tab()
        
        # 底部状态栏
        self.status_bar = ttk.Label(
            main_frame, 
            text="就绪", 
            bootstyle="secondary-inverse",
            font=("Microsoft YaHei UI", 9)
        )
        self.status_bar.pack(fill=X, pady=(10, 0))
    
    def change_theme(self, event=None):
        """更改主题"""
        theme = self.theme_var.get()
        self.root.style.theme_use(theme)
        self.update_status(f"主题已切换为: {theme}")
    
    def update_status(self, message, duration=3000):
        """更新状态栏"""
        self.status_bar.config(text=message)
        if duration > 0:
            self.root.after(duration, lambda: self.status_bar.config(text="就绪"))
        
    def create_custom_add_tab(self):
        """创建自定义新增选项卡"""
        tab = ttk.Frame(self.notebook, padding=20)
        self.notebook.add(tab, text="📝 自定义新增")
        
        # 创建卡片框架
        card_frame = ttk.Frame(tab, bootstyle="light")
        card_frame.pack(fill=BOTH, expand=YES)
        
        # 输入区域
        input_frame = ttk.Labelframe(card_frame, text="输入信息", padding=20, bootstyle="primary")
        input_frame.pack(fill=X, padx=20, pady=20)
        
        # 激活码输入
        ttk.Label(input_frame, text="激活码:", font=("Microsoft YaHei UI", 11, "bold")).grid(
            row=0, column=0, sticky=W, pady=15, padx=10)
        self.custom_key_entry = ttk.Entry(input_frame, width=40, font=("Consolas", 11), bootstyle="primary")
        self.custom_key_entry.grid(row=0, column=1, pady=15, padx=10, sticky=EW)
        
        # 产品名输入
        ttk.Label(input_frame, text="产品名:", font=("Microsoft YaHei UI", 11, "bold")).grid(
            row=1, column=0, sticky=W, pady=15, padx=10)
        self.custom_name_entry = ttk.Entry(input_frame, width=40, font=("Microsoft YaHei UI", 11), bootstyle="primary")
        self.custom_name_entry.grid(row=1, column=1, pady=15, padx=10, sticky=EW)
        
        input_frame.columnconfigure(1, weight=1)
        
        # 按钮区域
        btn_frame = ttk.Frame(input_frame)
        btn_frame.grid(row=2, column=0, columnspan=2, pady=20)
        
        ttk.Button(
            btn_frame, 
            text="✓ 添加到数据库", 
            command=self.custom_add_product,
            bootstyle="success",
            width=15
        ).pack(side=LEFT, padx=10)
        
        ttk.Button(
            btn_frame, 
            text="✗ 清空", 
            command=self.clear_custom_fields,
            bootstyle="warning",
            width=15
        ).pack(side=LEFT, padx=10)
        
        # 结果显示区域
        result_frame = ttk.Labelframe(card_frame, text="操作结果", padding=20, bootstyle="info")
        result_frame.pack(fill=BOTH, expand=YES, padx=20, pady=(0, 20))
        
        self.custom_result = scrolledtext.ScrolledText(
            result_frame, 
            width=70, 
            height=15, 
            font=("Consolas", 10),
            wrap=tk.WORD
        )
        self.custom_result.pack(fill=BOTH, expand=YES)
        
    def create_batch_add_tab(self):
        """创建批量新增选项卡"""
        tab = ttk.Frame(self.notebook, padding=20)
        self.notebook.add(tab, text="📦 批量新增")
        
        # 创建卡片框架
        card_frame = ttk.Frame(tab, bootstyle="light")
        card_frame.pack(fill=BOTH, expand=YES)
        
        # 输入区域
        input_frame = ttk.Labelframe(card_frame, text="批量生成设置", padding=20, bootstyle="success")
        input_frame.pack(fill=X, padx=20, pady=20)
        
        # 产品名输入
        ttk.Label(input_frame, text="产品名:", font=("Microsoft YaHei UI", 11, "bold")).grid(
            row=0, column=0, sticky=W, pady=15, padx=10)
        self.batch_name_entry = ttk.Entry(input_frame, width=40, font=("Microsoft YaHei UI", 11), bootstyle="success")
        self.batch_name_entry.grid(row=0, column=1, pady=15, padx=10, sticky=EW)
        
        # 生成数量输入
        ttk.Label(input_frame, text="生成数量:", font=("Microsoft YaHei UI", 11, "bold")).grid(
            row=1, column=0, sticky=W, pady=15, padx=10)
        self.batch_count_entry = ttk.Entry(input_frame, width=40, font=("Microsoft YaHei UI", 11), bootstyle="success")
        self.batch_count_entry.grid(row=1, column=1, pady=15, padx=10, sticky=EW)
        self.batch_count_entry.insert(0, "10")
        
        input_frame.columnconfigure(1, weight=1)
        
        # 进度条
        progress_frame = ttk.Frame(input_frame)
        progress_frame.grid(row=2, column=0, columnspan=2, pady=15, sticky=EW)
        
        self.progress_bar = ttk.Progressbar(
            progress_frame, 
            bootstyle="success-striped",
            mode='determinate'
        )
        self.progress_bar.pack(fill=X, padx=10)
        
        # 按钮区域
        btn_frame = ttk.Frame(input_frame)
        btn_frame.grid(row=3, column=0, columnspan=2, pady=20)
        
        self.batch_add_btn = ttk.Button(
            btn_frame, 
            text="🚀 批量生成并添加", 
            command=self.batch_add_products,
            bootstyle="success",
            width=20
        )
        self.batch_add_btn.pack(side=LEFT, padx=10)
        
        ttk.Button(
            btn_frame, 
            text="✗ 清空", 
            command=self.clear_batch_fields,
            bootstyle="warning",
            width=15
        ).pack(side=LEFT, padx=10)
        
        # 结果显示区域
        result_frame = ttk.Labelframe(card_frame, text="批量生成日志", padding=20, bootstyle="info")
        result_frame.pack(fill=BOTH, expand=YES, padx=20, pady=(0, 20))
        
        self.batch_result = scrolledtext.ScrolledText(
            result_frame, 
            width=70, 
            height=15, 
            font=("Consolas", 10),
            wrap=tk.WORD
        )
        self.batch_result.pack(fill=BOTH, expand=YES)
        
    def create_query_tab(self):
        """创建查询选项卡"""
        tab = ttk.Frame(self.notebook, padding=20)
        self.notebook.add(tab, text="🔍 查询")
        
        # 创建卡片框架
        card_frame = ttk.Frame(tab, bootstyle="light")
        card_frame.pack(fill=BOTH, expand=YES)
        
        # 查询区域
        search_frame = ttk.Labelframe(card_frame, text="查询条件", padding=20, bootstyle="info")
        search_frame.pack(fill=X, padx=20, pady=20)
        
        query_input_frame = ttk.Frame(search_frame)
        query_input_frame.pack(fill=X)
        
        ttk.Label(query_input_frame, text="产品名:", font=("Microsoft YaHei UI", 11, "bold")).pack(
            side=LEFT, padx=(0, 10))
        
        self.query_name_entry = ttk.Entry(
            query_input_frame, 
            width=30, 
            font=("Microsoft YaHei UI", 11),
            bootstyle="info"
        )
        self.query_name_entry.pack(side=LEFT, padx=10, fill=X, expand=YES)
        
        ttk.Button(
            query_input_frame, 
            text="🔍 查询", 
            command=self.query_by_product_name,
            bootstyle="info",
            width=12
        ).pack(side=LEFT, padx=5)
        
        ttk.Button(
            query_input_frame, 
            text="🗑 清空结果", 
            command=self.clear_query_result,
            bootstyle="secondary",
            width=12
        ).pack(side=LEFT, padx=5)
        
        ttk.Button(
            query_input_frame, 
            text="📊 导出Excel", 
            command=self.export_to_excel,
            bootstyle="success",
            width=12
        ).pack(side=LEFT, padx=5)
        
        # 结果统计
        self.query_stats_label = ttk.Label(
            search_frame, 
            text="查询结果: 0 条记录", 
            font=("Microsoft YaHei UI", 10, "bold"),
            bootstyle="info"
        )
        self.query_stats_label.pack(anchor=W, pady=(15, 0))
        
        # 结果显示区域
        result_frame = ttk.Labelframe(card_frame, text="查询结果", padding=20, bootstyle="primary")
        result_frame.pack(fill=BOTH, expand=YES, padx=20, pady=(0, 20))
        
        # 创建Treeview
        tree_frame = ttk.Frame(result_frame)
        tree_frame.pack(fill=BOTH, expand=YES)
        
        # 定义列
        columns = ("激活码", "产品名", "状态", "绑定码", "创建时间", "使用时间")
        self.query_tree = ttk.Treeview(
            tree_frame, 
            columns=columns, 
            show="headings", 
            height=15,
            bootstyle="info"
        )
        
        # 设置列
        self.query_tree.heading("激活码", text="激活码")
        self.query_tree.heading("产品名", text="产品名")
        self.query_tree.heading("状态", text="状态")
        self.query_tree.heading("绑定码", text="绑定码")
        self.query_tree.heading("创建时间", text="创建时间")
        self.query_tree.heading("使用时间", text="使用时间")
        
        # 设置列宽
        self.query_tree.column("激活码", width=150, anchor=W)
        self.query_tree.column("产品名", width=120, anchor=CENTER)
        self.query_tree.column("状态", width=80, anchor=CENTER)
        self.query_tree.column("绑定码", width=150, anchor=W)
        self.query_tree.column("创建时间", width=150, anchor=CENTER)
        self.query_tree.column("使用时间", width=150, anchor=CENTER)
        
        # 添加滚动条
        scrollbar = ttk.Scrollbar(tree_frame, orient=VERTICAL, command=self.query_tree.yview)
        self.query_tree.configure(yscrollcommand=scrollbar.set)
        
        # 布局
        self.query_tree.pack(side=LEFT, fill=BOTH, expand=YES)
        scrollbar.pack(side=RIGHT, fill=Y)
        
        # 创建右键菜单
        self.context_menu = tk.Menu(self.query_tree, tearoff=0)
        self.context_menu.add_command(label="📋 复制激活码", command=self.copy_product_key)
        self.context_menu.add_command(label="📋 复制绑定码", command=self.copy_response_key)
        self.context_menu.add_separator()
        self.context_menu.add_command(label="📄 复制整行数据", command=self.copy_row_data)
        
        # 绑定事件
        self.query_tree.bind("<Button-3>", self.show_context_menu)
        self.query_tree.bind("<Double-Button-1>", self.on_double_click)
        
    def generate_random_key(self, length=16):
        """生成随机激活码"""
        characters = string.ascii_letters + string.digits
        return ''.join(random.choice(characters) for _ in range(length))
    
    def custom_add_product(self):
        """自定义新增产品"""
        product_key = self.custom_key_entry.get().strip()
        product_name = self.custom_name_entry.get().strip()
        
        if not product_key or not product_name:
            messagebox.showwarning("警告", "请输入激活码和产品名！", parent=self.root)
            return
        
        self.custom_result.insert(tk.END, f"正在添加: {product_key} - {product_name}\n")
        self.custom_result.see(tk.END)
        self.update_status(f"正在添加产品: {product_name}")
        
        try:
            response = requests.post(f"{BASE_URL}/add_product", json={
                "product_key": product_key,
                "product_name": product_name
            }, timeout=5)
            
            if response.status_code == 200:
                result = response.json()
                self.custom_result.insert(tk.END, 
                    f"✓ 成功: {json.dumps(result, ensure_ascii=False)}\n\n")
                self.custom_result.see(tk.END)
                self.update_status("✓ 产品添加成功！")
                messagebox.showinfo("成功", "产品添加成功！", parent=self.root)
            else:
                self.custom_result.insert(tk.END, 
                    f"✗ 失败 (状态码: {response.status_code}): {response.text}\n\n")
                self.custom_result.see(tk.END)
                self.update_status("✗ 添加失败", 0)
                messagebox.showerror("错误", f"添加失败: {response.text}", parent=self.root)
                
        except requests.exceptions.RequestException as e:
            error_msg = f"✗ 连接错误: {str(e)}\n\n"
            self.custom_result.insert(tk.END, error_msg)
            self.custom_result.see(tk.END)
            self.update_status("✗ 连接错误", 0)
            messagebox.showerror("连接错误", f"无法连接到服务器: {str(e)}", parent=self.root)
    
    def batch_add_products(self):
        """批量新增产品"""
        product_name = self.batch_name_entry.get().strip()
        count_str = self.batch_count_entry.get().strip()
        
        if not product_name:
            messagebox.showwarning("警告", "请输入产品名！", parent=self.root)
            return
        
        try:
            count = int(count_str)
            if count <= 0:
                raise ValueError("数量必须大于0")
            if count > 10000:
                if not messagebox.askyesno("确认", 
                    f"您要生成 {count} 条记录，这可能需要较长时间，确定继续吗？", parent=self.root):
                    return
        except ValueError:
            messagebox.showwarning("警告", f"请输入有效的数量（正整数）！", parent=self.root)
            return
        
        # 在新线程中执行批量添加
        thread = threading.Thread(target=self._batch_add_worker, 
                                 args=(product_name, count))
        thread.daemon = True
        thread.start()
    
    def _batch_add_worker(self, product_name, count):
        """批量添加的工作线程"""
        self.batch_add_btn.config(state='disabled')
        
        self.batch_result.insert(tk.END, f"开始批量生成 {count} 个激活码...\n")
        self.batch_result.insert(tk.END, f"产品名: {product_name}\n")
        self.batch_result.insert(tk.END, "=" * 60 + "\n")
        self.batch_result.see(tk.END)
        
        success_count = 0
        fail_count = 0
        
        for i in range(count):
            product_key = self.generate_random_key(16)
            
            try:
                response = requests.post(f"{BASE_URL}/add_product", json={
                    "product_key": product_key,
                    "product_name": product_name
                }, timeout=5)
                
                if response.status_code == 200:
                    success_count += 1
                    if i < 10 or (i + 1) % 100 == 0:
                        self.batch_result.insert(tk.END, 
                            f"[{i+1}/{count}] ✓ {product_key}\n")
                        self.batch_result.see(tk.END)
                else:
                    fail_count += 1
                    self.batch_result.insert(tk.END, 
                        f"[{i+1}/{count}] ✗ {product_key} - 失败: {response.text}\n")
                    self.batch_result.see(tk.END)
                    
            except requests.exceptions.RequestException as e:
                fail_count += 1
                self.batch_result.insert(tk.END, 
                    f"[{i+1}/{count}] ✗ {product_key} - 连接错误: {str(e)}\n")
                self.batch_result.see(tk.END)
            
            # 更新进度条
            progress = (i + 1) / count * 100
            self.progress_bar['value'] = progress
            self.root.update_idletasks()
            self.update_status(f"正在生成... {i+1}/{count} ({progress:.1f}%)", 0)
        
        # 显示总结
        self.batch_result.insert(tk.END, "\n" + "=" * 60 + "\n")
        self.batch_result.insert(tk.END, f"批量添加完成！\n")
        self.batch_result.insert(tk.END, f"总计: {count} 条\n")
        self.batch_result.insert(tk.END, f"成功: {success_count} 条\n")
        self.batch_result.insert(tk.END, f"失败: {fail_count} 条\n")
        self.batch_result.insert(tk.END, "=" * 60 + "\n\n")
        self.batch_result.see(tk.END)
        
        # 重置进度条
        self.progress_bar['value'] = 0
        
        # 启用按钮
        self.batch_add_btn.config(state='normal')
        
        # 更新状态
        self.update_status(f"✓ 批量添加完成！成功: {success_count}, 失败: {fail_count}")
        
        messagebox.showinfo("完成", 
            f"批量添加完成！\n成功: {success_count} 条\n失败: {fail_count} 条", parent=self.root)
    
    def query_by_product_name(self):
        """根据产品名查询"""
        product_name = self.query_name_entry.get().strip()
        
        if not product_name:
            messagebox.showwarning("警告", "请输入产品名！", parent=self.root)
            return
        
        self.update_status(f"正在查询产品: {product_name}...", 0)
        
        try:
            response = requests.get(f"{BASE_URL}/query_by_product_name", 
                                   params={"product_name": product_name}, 
                                   timeout=5)
            
            if response.status_code == 200:
                result = response.json()
                
                # 清空之前的结果
                for item in self.query_tree.get_children():
                    self.query_tree.delete(item)
                
                # 检查结果格式
                if isinstance(result, dict) and "codes" in result:
                    data_list = result["codes"]
                elif isinstance(result, dict) and "data" in result:
                    data_list = result["data"]
                elif isinstance(result, list):
                    data_list = result
                else:
                    data_list = []
                
                # 填充表格
                for item in data_list:
                    if isinstance(item, dict):
                        key = item.get("product_key", "")
                        name = item.get("product_name", "")
                        is_used = item.get("is_used", None)
                        if is_used is True:
                            status = "已使用"
                        elif is_used is False:
                            status = "未使用"
                        else:
                            status = item.get("status", "未知")
                        response_key = item.get("response_key", "") if item.get("response_key") else ""
                        created = item.get("created_at", "")
                        used = item.get("used_at", "") if item.get("used_at") else ""
                    else:
                        key = str(item)
                        name = product_name
                        status = "未知"
                        response_key = ""
                        created = ""
                        used = ""
                    
                    # 根据状态设置不同的标签
                    tag = "used" if status == "已使用" else "unused"
                    self.query_tree.insert("", tk.END, values=(key, name, status, 
                                                              response_key, created, used),
                                         tags=(tag,))
                
                # 配置标签颜色
                self.query_tree.tag_configure("used", foreground="#dc3545")
                self.query_tree.tag_configure("unused", foreground="#28a745")
                
                # 更新统计信息
                count = len(data_list)
                self.query_stats_label.config(text=f"查询结果: {count} 条记录")
                
                if count == 0:
                    self.update_status(f"未找到产品 '{product_name}' 的相关记录")
                    messagebox.showinfo("提示", f"未找到产品 '{product_name}' 的相关记录", parent=self.root)
                else:
                    self.update_status(f"✓ 查询成功！找到 {count} 条记录")
                    messagebox.showinfo("成功", f"查询成功！共找到 {count} 条记录", parent=self.root)
            else:
                self.update_status("✗ 查询失败", 0)
                messagebox.showerror("错误", 
                    f"查询失败 (状态码: {response.status_code}): {response.text}", parent=self.root)
                
        except requests.exceptions.RequestException as e:
            self.update_status("✗ 连接错误", 0)
            messagebox.showerror("连接错误", f"无法连接到服务器: {str(e)}", parent=self.root)
    
    def clear_custom_fields(self):
        """清空自定义新增字段"""
        self.custom_key_entry.delete(0, tk.END)
        self.custom_name_entry.delete(0, tk.END)
        self.update_status("已清空输入字段")
    
    def clear_batch_fields(self):
        """清空批量新增字段"""
        self.batch_name_entry.delete(0, tk.END)
        self.batch_count_entry.delete(0, tk.END)
        self.batch_count_entry.insert(0, "10")
        self.batch_result.delete(1.0, tk.END)
        self.progress_bar['value'] = 0
        self.update_status("已清空输入字段")
    
    def clear_query_result(self):
        """清空查询结果"""
        for item in self.query_tree.get_children():
            self.query_tree.delete(item)
        self.query_stats_label.config(text="查询结果: 0 条记录")
        self.update_status("已清空查询结果")
    
    def show_context_menu(self, event):
        """显示右键菜单"""
        item = self.query_tree.identify_row(event.y)
        if item:
            self.query_tree.selection_set(item)
            self.context_menu.post(event.x_root, event.y_root)
    
    def copy_product_key(self):
        """复制激活码到剪贴板"""
        selection = self.query_tree.selection()
        if selection:
            item = selection[0]
            values = self.query_tree.item(item, "values")
            if values and len(values) > 0:
                product_key = values[0]
                self.root.clipboard_clear()
                self.root.clipboard_append(product_key)
                self.root.update()
                self.update_status(f"✓ 已复制激活码: {product_key}")
    
    def copy_response_key(self):
        """复制绑定码到剪贴板"""
        selection = self.query_tree.selection()
        if selection:
            item = selection[0]
            values = self.query_tree.item(item, "values")
            if values and len(values) > 3:
                response_key = values[3]
                if response_key:
                    self.root.clipboard_clear()
                    self.root.clipboard_append(response_key)
                    self.root.update()
                    self.update_status(f"✓ 已复制绑定码: {response_key}")
                else:
                    self.update_status("⚠ 该记录没有绑定码")
    
    def copy_row_data(self):
        """复制整行数据到剪贴板"""
        selection = self.query_tree.selection()
        if selection:
            item = selection[0]
            values = self.query_tree.item(item, "values")
            if values:
                row_data = f"激活码: {values[0]}\n产品名: {values[1]}\n状态: {values[2]}\n绑定码: {values[3]}\n创建时间: {values[4]}\n使用时间: {values[5]}"
                self.root.clipboard_clear()
                self.root.clipboard_append(row_data)
                self.root.update()
                self.update_status("✓ 已复制整行数据")
    
    def on_double_click(self, event):
        """双击事件处理"""
        region = self.query_tree.identify_region(event.x, event.y)
        if region == "cell":
            column = self.query_tree.identify_column(event.x)
            item = self.query_tree.identify_row(event.y)
            if item:
                values = self.query_tree.item(item, "values")
                col_index = int(column.replace("#", "")) - 1
                
                if col_index == 0:  # 激活码列
                    product_key = values[0]
                    self.root.clipboard_clear()
                    self.root.clipboard_append(product_key)
                    self.root.update()
                    self.update_status(f"✓ 已复制激活码: {product_key}")
                
                elif col_index == 3:  # 绑定码列
                    response_key = values[3]
                    if response_key:
                        self.root.clipboard_clear()
                        self.root.clipboard_append(response_key)
                        self.root.update()
                        self.update_status(f"✓ 已复制绑定码: {response_key}")
                    else:
                        self.update_status("⚠ 该记录没有绑定码")
    
    def export_to_excel(self):
        """导出查询结果到Excel"""
        # 检查是否有数据
        items = self.query_tree.get_children()
        if not items:
            messagebox.showwarning("警告", "没有数据可以导出！\n请先查询数据。", parent=self.root)
            return
        
        # 选择保存位置
        default_filename = f"查询结果_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        file_path = filedialog.asksaveasfilename(
            parent=self.root,
            title="保存Excel文件",
            defaultextension=".xlsx",
            initialfile=default_filename,
            filetypes=[("Excel文件", "*.xlsx"), ("所有文件", "*.*")]
        )
        
        if not file_path:
            return
        
        self.update_status("正在导出Excel文件...", 0)
        
        try:
            # 尝试使用 openpyxl
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, Alignment, PatternFill
                
                # 创建工作簿
                wb = Workbook()
                ws = wb.active
                ws.title = "查询结果"
                
                # 写入表头
                headers = ["激活码", "产品名", "状态", "绑定码", "创建时间", "使用时间"]
                for col, header in enumerate(headers, start=1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = Font(bold=True, size=12)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                    cell.font = Font(bold=True, size=12, color="FFFFFF")
                
                # 写入数据
                for row_idx, item in enumerate(items, start=2):
                    values = self.query_tree.item(item, "values")
                    for col_idx, value in enumerate(values, start=1):
                        cell = ws.cell(row=row_idx, column=col_idx, value=str(value))
                        cell.alignment = Alignment(horizontal='left', vertical='center')
                        
                        # 根据状态设置颜色
                        if col_idx == 3:  # 状态列
                            if value == "已使用":
                                cell.font = Font(color="DC3545")
                            elif value == "未使用":
                                cell.font = Font(color="28A745")
                
                # 设置列宽
                column_widths = [25, 20, 12, 25, 25, 25]
                for col, width in enumerate(column_widths, start=1):
                    ws.column_dimensions[chr(64 + col)].width = width
                
                # 保存文件
                wb.save(file_path)
                
                count = len(items)
                self.update_status(f"✓ 成功导出 {count} 条记录到 Excel")
                messagebox.showinfo("成功", 
                    f"已成功导出 {count} 条记录！\n\n文件保存至:\n{file_path}", 
                    parent=self.root)
                
            except ImportError:
                # 如果没有 openpyxl，使用 CSV 作为替代
                import csv
                
                csv_path = file_path.replace('.xlsx', '.csv')
                
                with open(csv_path, 'w', newline='', encoding='utf-8-sig') as f:
                    writer = csv.writer(f)
                    
                    # 写入表头
                    writer.writerow(["激活码", "产品名", "状态", "绑定码", "创建时间", "使用时间"])
                    
                    # 写入数据
                    for item in items:
                        values = self.query_tree.item(item, "values")
                        writer.writerow(values)
                
                count = len(items)
                self.update_status(f"✓ 成功导出 {count} 条记录到 CSV")
                messagebox.showinfo("提示", 
                    f"已成功导出 {count} 条记录到 CSV 格式！\n\n" +
                    "（未安装 openpyxl，无法导出 Excel 格式）\n" +
                    f"文件保存至:\n{csv_path}\n\n" +
                    "若需要 Excel 格式，请安装: pip install openpyxl", 
                    parent=self.root)
                
        except Exception as e:
            self.update_status(f"✗ 导出失败: {str(e)}", 0)
            messagebox.showerror("错误", f"导出失败：{str(e)}", parent=self.root)


def main():
    # 创建主窗口，可以选择不同的主题
    # 亮色主题: 'cosmo', 'flatly', 'litera', 'minty', 'lumen', 'sandstone', 'yeti', 'pulse', 'united', 'morph', 'journal', 'simplex', 'cerculean'
    # 暗色主题: 'darkly', 'superhero', 'solar', 'cyborg', 'vapor'
    
    # 从配置文件读取默认主题
    default_theme = CONFIG.get('APP', 'default_theme', fallback='cosmo')
    root = ttk.Window(themename=default_theme)
    
    # 从配置文件读取窗口大小
    window_width = CONFIG.getint('APP', 'window_width', fallback=1000)
    window_height = CONFIG.getint('APP', 'window_height', fallback=750)
    root.geometry(f"{window_width}x{window_height}")
    
    app = ModernAdminApp(root)
    
    # 在状态栏显示当前配置
    app.update_status(f"后端服务器: {BASE_URL}", 5000)
    
    root.mainloop()


if __name__ == "__main__":
    main()

